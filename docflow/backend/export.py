"""File exporters for JSON / CSV / XLSX / XML."""
from __future__ import annotations

import io
import json
import re
from datetime import datetime
from typing import Any

from models import DocumentRecord, ExportFormat


_FORBIDDEN_CHARS = re.compile(r'[\\/:*?"<>|]+')


def sanitize_filename(name: str) -> str:
    cleaned = _FORBIDDEN_CHARS.sub("", name or "").strip()
    cleaned = re.sub(r"\s+", "_", cleaned)
    return cleaned or "document"


def get_download_filename(record: DocumentRecord, fmt: ExportFormat) -> str:
    ext = fmt.value
    number = record.data.document_number if record.data else ""
    if number:
        return f"{sanitize_filename(number)}.{ext}"
    return f"unknown_{record.doc_id[:8]}.{ext}"


def _payload_dict(record: DocumentRecord) -> dict[str, Any]:
    data = record.data.model_dump() if record.data else {}
    return {
        "doc_id": record.doc_id,
        "exported_at": datetime.utcnow().isoformat(timespec="seconds"),
        "document_type": data.get("document_type", ""),
        "document_number": data.get("document_number", ""),
        "document_date": data.get("document_date", ""),
        "supplier_name": data.get("supplier_name", ""),
        "supplier_inn": data.get("supplier_inn", ""),
        "buyer_name": data.get("buyer_name", ""),
        "buyer_inn": data.get("buyer_inn", ""),
        "total_amount": data.get("total_amount", ""),
        "vat_amount": data.get("vat_amount", ""),
        "currency": data.get("currency", "RUB"),
        "items": data.get("items", []),
    }


def export_json(record: DocumentRecord) -> bytes:
    payload = _payload_dict(record)
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def export_csv(record: DocumentRecord) -> bytes:
    try:
        import pandas as pd  # type: ignore
    except Exception:  # pragma: no cover - fallback path
        return _export_csv_fallback(record)

    payload = _payload_dict(record)
    items = payload.pop("items") or []
    header_df = pd.DataFrame(
        [(k, v) for k, v in payload.items() if k != "items"],
        columns=["Поле", "Значение"],
    )
    items_df = pd.DataFrame(
        items or [],
        columns=["name", "quantity", "unit", "unit_price", "total_price"],
    )
    items_df = items_df.rename(
        columns={
            "name": "Наименование",
            "quantity": "Количество",
            "unit": "Ед.изм.",
            "unit_price": "Цена",
            "total_price": "Сумма",
        }
    )

    buffer = io.StringIO()
    header_df.to_csv(buffer, index=False)
    buffer.write("\n")
    buffer.write("Товары\n")
    items_df.to_csv(buffer, index=False)
    # UTF-8 with BOM for Excel
    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


def _export_csv_fallback(record: DocumentRecord) -> bytes:
    import csv

    payload = _payload_dict(record)
    items = payload.pop("items") or []
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Поле", "Значение"])
    for k, v in payload.items():
        writer.writerow([k, v])
    writer.writerow([])
    writer.writerow(["Товары"])
    writer.writerow(["Наименование", "Количество", "Ед.изм.", "Цена", "Сумма"])
    for item in items:
        writer.writerow(
            [
                item.get("name", ""),
                item.get("quantity", ""),
                item.get("unit", ""),
                item.get("unit_price", ""),
                item.get("total_price", ""),
            ]
        )
    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


def export_xlsx(record: DocumentRecord) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    payload = _payload_dict(record)
    items = payload.pop("items") or []

    wb = Workbook()
    ws_req = wb.active
    ws_req.title = "Реквизиты"
    ws_req.append(["Поле", "Значение"])
    ws_req["A1"].font = Font(bold=True)
    ws_req["B1"].font = Font(bold=True)
    for key, value in payload.items():
        ws_req.append([key, value])
    ws_req.column_dimensions["A"].width = 30
    ws_req.column_dimensions["B"].width = 60

    ws_items = wb.create_sheet("Товары")
    headers = ["Наименование", "Кол-во", "Ед.", "Цена", "Сумма"]
    ws_items.append(headers)
    for col_idx in range(1, len(headers) + 1):
        ws_items.cell(row=1, column=col_idx).font = Font(bold=True)
    for item in items:
        ws_items.append(
            [
                item.get("name", ""),
                item.get("quantity", ""),
                item.get("unit", ""),
                item.get("unit_price", ""),
                item.get("total_price", ""),
            ]
        )
    ws_items.column_dimensions["A"].width = 40
    ws_items.column_dimensions["B"].width = 12
    ws_items.column_dimensions["C"].width = 10
    ws_items.column_dimensions["D"].width = 14
    ws_items.column_dimensions["E"].width = 14

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_xml(record: DocumentRecord) -> bytes:
    payload = _payload_dict(record)
    try:
        from dicttoxml import dicttoxml  # type: ignore

        xml_bytes = dicttoxml(
            payload,
            custom_root="document",
            attr_type=False,
        )
        return xml_bytes
    except Exception:
        return _export_xml_fallback(payload)


def _export_xml_fallback(payload: dict[str, Any]) -> bytes:
    import xml.etree.ElementTree as ET

    root = ET.Element("document")
    for key, value in payload.items():
        if key == "items":
            items_el = ET.SubElement(root, "items")
            for item in value or []:
                item_el = ET.SubElement(items_el, "item")
                for k, v in item.items():
                    child = ET.SubElement(item_el, k)
                    child.text = str(v)
            continue
        child = ET.SubElement(root, key)
        child.text = "" if value is None else str(value)
    xml_bytes = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    return xml_bytes


def export(record: DocumentRecord, fmt: ExportFormat) -> tuple[bytes, str, str]:
    """Return (bytes, media_type, download_filename)."""
    filename = get_download_filename(record, fmt)
    if fmt is ExportFormat.json:
        return export_json(record), "application/json", filename
    if fmt is ExportFormat.csv:
        return export_csv(record), "text/csv; charset=utf-8", filename
    if fmt is ExportFormat.xlsx:
        return (
            export_xlsx(record),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename,
        )
    if fmt is ExportFormat.xml:
        return export_xml(record), "application/xml; charset=utf-8", filename
    raise ValueError(f"Unsupported format: {fmt}")
